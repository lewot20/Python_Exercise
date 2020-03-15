
###Print 연습하기
# print('What is your name?') #ask for their name
# MyName = input()
# print('It is good to meet you, ' + MyName)
# print ('The length of your name is :')
# print(len(MyName))
# print('What is your age?')
# MyAge = input()
# print('You will be ' +str(int(MyAge) +1) + ' in a year')

###while 문 만들기
# name =''
# while name != 'your name':
#     print('Please type your name.')
#     name= input()
# print('Thank you!')

###break 문 만들기
# while True:
#     print('Please type your name.')
#     name = input()
#     if name == 'your name':
#         break
# print('Thank you!')

###continue 문 만들기
# while True:
#     print('Who are you?')
#     name = input()
#     if name != 'Joe':
#         continue
#     print('Hello Joe. What is the password? (It is a fish.)')
#     password = input()
#     if password =='swordfish':
#         break
# print('Access granted')

###  0은 False이다 빈칸''도 False이다
# name = ''
# while not name :
#     print('Enter your name:')
#     name = input()
# print('How many guests will you have?')
# num0fGuests = int(input())
# if num0fGuests:
#     print('Be sure to have enough room for all your guests.')
# print('Done')

# ### For 루프와 range() 함수 연습하기
#
# print('My name is')
# for i in range(5):
#     print('Jimmy Five Times (' + str(i) + ')')
#
#
# ### import 연습하기
# ### import random, sys, os, math
#
#
# from random import *
# for i in range(5):
#     print(randint(1,10))
#
#
# import sys
# while True:
#     print('type exit to exit.')
#     response = input()
#     if response == 'exit':
#         sys.exit()
#     print( 'You typed ' + response + '.')

# ##함수 만들기
# def hello():
#     print('what is your name?')
#     name = input()
#     print('Hello ' + str(name))
#
# hello()
#
# def spam(divideby):
#     try:
#         return 42/divideby
#     except ZeroDivisionError:
#         print('Error')
#
# print(spam(2))
# print(spam(0))
# print(spam(33))
#
# ## this is a guess the number game
# from random import *
# secretnumber = randint(1,20)
# print("I' am thinking of a number between 1 to 20" )
# print('Correct the answer in six times')
# for guesstaken in range(1,7):
#     print('Take a guess.')
#     guess = int(input())
#     if guess < secretnumber:
#         print('your guess is low.')
#     elif guess > secretnumber:
#         print('your guess is high')
#     else :
#         break #this condition is the correct guess!
#
# if guess  == secretnumber:
#     print('Good job! You guessed my number in ' + str(guesstaken) + ' guesses!')
# else:
#     print('Nope. The number i was thinking of was ' + str(secretnumber))
#
#
# ##collatz 수열 만들기
# print('Please Enter number:')
# number = input()
# print(type(int(number)))
# while True :
#     if str(type(int(number))) == "<class 'int'>":
#         print('Good')
#             continue
#     print('please input the integer')
#
# ###다시 collatz 수열 만들기
# def collatz(num):
#     answer = 0
#     while num!=1:
#         if num % 2 == 0 :
#             num = num/2
#             print(int(num))
#         else :
#             num = num * 3 + 1
#             print(int(num))
#         answer = answer +1
#     return(answer)
# print('Please input the number')
# num = input()
#
# def is_integer(x):
#     try:
#         return int(x)
#
#     except ValueError:
#         print('Please input the integer')
#
# print('the number of cycle is ' + str(collatz(is_integer(num))))

# ##리스트 만들기
# catnames=[]
# while True:
#     print('Enter the name of Cat ' + str(len(catnames) +1) + ' (Or enter nothing to stop)')
#     name = input()
#     if name =='':
#         break
#     catnames = catnames + [name]
# print('The cat names are:')
# for name in catnames:
#     print(' ' + name)
#
#
# ## 리스트와 함께 루프 사용하기
# supplies = ['pens','staplers','flame-throwers','blinders']
# for i in range(len(supplies)):
#     print('Index ' + str(i) + ' in supplies is : ' + supplies[i])
#
# ##index() 방법으로 리스트 안에서 값 찾기
# spam = ['hello','hi','howdy','heyas']
# spam.index('hello')
# spam.index('heyas')
#
# import copy
# spam = ['A','B','C','D']
# cheese = copy.copy(spam)
# cheese[1]= 42
# print(spam)
# print(cheese)
#
# ### 챕터4 프로그램예제
# def yoso(list):
#     for i in range(0, len(list)):
#         if i < len(list) - 1 :
#             print(list[i], end=',')
#         else :
#             print(' and ' + list[i], end='\n')
#
# spam = ['hello','hi','howdy','heyas']
# yoso(spam)
#
# ## dictionary 구조로 data 저장하기 (아직 서버는 없음)
# birthdays = {'Alice':'Apr 1', 'Bob':'Dec 12', 'Carol' :"Mar 4"}
# while True:
#     print('Enter a name: (blank to quit)')
#     name = input()
#     if name == '':
#         break
#     if name in birthdays:
#         print(birthdays[name] + ' is the birthday of ' + name)
#     else :
#         print('I do not have birthday information')
#         print('What is their birthday?')
#         print('if you want to quit, please enter')
#         bday = input()
#         if bday == '':
#             break
#         birthdays[name] = bday
#         print('Birthday database updated.')
#
# ### 틱택토 보드 게임 만들기
# theboard = {'top-L':' ','top-M' :' ','top-R':' ','mid-L':' ','mid-M':' ','mid-R':' ','low-L':' ','low-M':' ','low-R':' '}
# def printBoard(board):
#     turn = 'X'
#     for i in range(9):
#         print(board['top-L'] + '|' + board['top-M'] + '|' + board['top-R'])
#         print('-+-+-')
#         print(board['mid-L'] + '|' + board['mid-M'] + '|' + board['mid-R'])
#         print('-+-+-')
#         print(board['low-L'] + '|' + board['low-M'] + '|' + board['low-R'])
#
#
#         print('Turn for ' + turn +'. Move on which space?')
#         print("'top, mid, low + - + L, M, R'")
#         move = input()
#         theboard[move] = turn
#         if turn =='X':
#             turn ='O'
#         else :
#             turn = 'X'
#
# printBoard(theboard)

# ## 숫자인지 문자인지 검정하기
# while True:
#     print('Enter your age')
#     age = input()
#     if age.isdecimal():
#         break
#     print('Please enter a number for your age')
#
# ## 정렬하기
# def printpicnic (itemdict, leftwidth, rigthwidth):
#     print('PICNIC ITEMS'.center(leftwidth+rigthwidth, '-'))
#     for k, v in itemdict.items():
#         print(k.ljust(leftwidth, '.') + str(v).rjust(rigthwidth))
#
# picnicitems={'sandwiches':3, 'apples': 12, 'cups':4, 'cookies':4300}
# printpicnic(picnicitems, 12,2)
# printpicnic(picnicitems, 20, 6)
#
# ### 공백 없애기
# spam = 'SpamSpamSpamEggsBaconSpamEggsSpam'
# print(spam.strip('mpaS'))
#
# ##pyperclip 모듈 사용하여 문자열 복사 또는 붙여넣기
# import pyperclip
# pyperclip.copy('Helloworld')
# print(pyperclip.paste())
#
# 프로그램 설계
# pw.py 다시 만들어봤지만 실행이 되지 않음
#
# ## 정규표현식 없이 텍스트 패턴 찾기(p176)
# def isPhoneNumber(text):
#     if len(text) != 12:
#         return False
#     for i in range(0,3):
#         if not text[i].isdecimal():
#             return False
#     if text[3] != '-':
#         return False
#     for i in range(4,7):
#         if not text[i].isdecimal():
#             return False
#     if text[7] != '-':
#         return False
#     for i in range(8,12):
#         if not text[i].isdecimal():
#             return False
#     return True
#
# message = 'Call me at 415-555-1011 tomorrow. 415-555-0000 is my office.'
# message_2 = 'Call me at 010-6396-3575'
# def gumsa(message):
#     print("I'll check the " + str(message))
#     for i in range(len(message)):
#         chunk = message[i:i+12]
#         if isPhoneNumber(chunk):
#             print('Phone number found: ' + chunk)
#     print('Done')
#
# print(gumsa(message))
# print(gumsa(message_2))
#
#
# ## 정규식 객체 만들기
# import re
#
# phoneNumRegex = re.compile(r'\d\d\d-\d\d\d-\d\d\d\d')
# mo = phoneNumRegex.search('My number is 415-555-4242 and my office is 415-323-2311-.')
# print('phone number found: '+ mo.group())
# import re
# text = '나의 번호는 415-555-4242이라고 해 그리고 425-333-2413도 있어'
# phoneNumRegex = re.compile(r'(\d\d\d)-(\d\d\d-\d\d\d\d)')
# mo = phoneNumRegex.search(text)
# print(mo.group(1))
# print(mo.group(2))
# print(mo.group(0))
# print(mo.group())
# print(phoneNumRegex.findall(text))
# print(mo.groups())
# Areacode, maninumber = mo.groups()
# print(Areacode)
#
# batRegex = re.compile(r'Bat(wo)?man')
# mo1 = batRegex.search('The adventure of Batman')
# print(mo1.group())
#
# mo2 = batRegex.search('The adventure of Batwoman')
# print(mo2.group())
# print(mo2.group(1))
#
# ###정규표현식 ^, $
# begin = re.compile(r'^Hello')
# mo22 = begin.search('Hello world!')
# print(mo22.group())
# print(begin.findall('Hello world!'))
#
# agentregex = re.compile(r'Agent (\w\w)\w*')
# print(agentregex.sub(r'\1****', 'Agent Alice told Agent Carol that Agent eve knew Agent Bobbb was a double target'))
#
#
#
# ## 정규 표현식을 이용하여 프로젝트 전화번호와 이메일 추출하기
# #1 클립보드로부터 텍스트 가져오기
# #2 텍스트에서 전화번호와 주소를 찾기
# #3 클립보드에 붙여 넣기

#! python3
#phoneandemail.py - finds phone numbers and email adresses on the clipboard
#
# import pyperclip
# import re
#
# phoneregex = re.compile(r'''(
# (\d{3}|\(\d{3})?         #area code
# (\s|-|\.)?                      #separator
# (\d{3})                         #first 3 digits
# (\s|-|\.)                       #separator
# (\d{4})                         #first 4 digits
# (\s*(ext|x|ext.)\s*(\d{2,5}))?  #extension
# )''', re.VERBOSE)
#
# #TODO : Create email regex
# emailregex = re.compile(r'''(
# [a-zA-Z0-9._%+-] +
# @
# [a-zA-Z0-9.-]+
# (\.[a-zA-Z]{2,4})
# )''', re.VERBOSE)
#
# #TODO : Find matches in clipboard text
# text = str(pyperclip.paste())
#
# matches=[]
# for groups in phoneregex.findall(text):
#     phonenum = '-'.join([groups[1], groups[3], groups[5]])
#     if groups[8] != '':
#         phonenum += ' x' + groups[8]
#     matches.append(phonenum)
# for groups in emailregex.findall(text):
#     matches.append(groups[0])
#
# # TODO : clipboard 로
# if len(matches) >0:
#     pyperclip.copy('\n'.join(matches))
#     print('Copied to clipboard')
#     print('\n'.join(matches))
# else:
#     print('No phone numbers or email address found.')
#
#
# #TODO : Copy results to the clip board

## 8장 파일 읽고 쓰기
#
# import os
# print(os.path.join('usr','bin','spam'))
# myfiles = ['accounts.txt','details.csv','invite.docx']
# for filename in myfiles:
#     print(os.path.join('C:\\Users\\asweigart', filename))
#
# ### 현재 작업 디렉토리
#
# print(os.getcwd())
# # os.makedirs('D:\\pycharm\\qt5\\walnut\\waffles')
#
# calcfilepath = 'D:\\pycharm\\qt5\\walnut\\waffles\\calc.exe'
# print(os.path.split(calcfilepath))
# print(calcfilepath.split(os.path.sep))
#
# ## 파일 오픈하기
# hellofile = open('D:\\pycharm\\qt5\\walnut\\waffles\\masit.txt')
# hellocontent = hellofile.read()
# print(hellocontent)
#
# sonnetfile = open('D:\\pycharm\\qt5\\walnut\\waffles\\sonnet29.txt')
# print(sonnetfile.readlines())
#
#
# ### 파일 쓰고 닫기
# baconfile = open('D:\\pycharm\\qt5\\walnut\\waffles\\bacon.txt','w')
# baconfile.write('Hello world1111111111!\n')
# baconfile.close()
# baconfile = open('D:\\pycharm\\qt5\\walnut\\waffles\\bacon.txt','a')
# baconfile.write('bacon is not a vegetable')
# baconfile.close()
# baconfile = open('D:\\pycharm\\qt5\\walnut\\waffles\\bacon.txt')
# content = baconfile.read()
# baconfile.close()
# print(content)
#
# import shelve
# shelffile = shelve.open('mydata')
# cats = ['Zophie','Pooka','Simon']
# shelffile['cats'] = cats
# shelffile.close()
# shelffile = shelve.open('mydata')
# print(type(shelffile))
# print(shelffile['cats'])
# shelffile.close()
#
# ## pprint.pformat으로 변수 저장하기
#
# import pprint, shelve
# cats = [{'name':'Zophie','desc':'chubby'},{'name':'Pooka', 'desc':'fluffy'}]
# pprint.pformat(cats)
# fileObj = open('myCats.py', 'w')
# fileObj.write('cats = ' + pprint.pformat(cats) + '\n')
# fileObj.close()
#
# import myCats
# print(myCats.cats)
# print(myCats.cats[0])
# print(myCats.cats[0]['name'])


### 챕터 8 프로젝트 : 퀴즈 파일 만들기

# import random, os
# capitals = {
#     'Alabama' : 'Montgomery', 'Alaska' : 'Juneau', 'Arizona': 'Phoenix' , 'Arkansas' : 	'Little Rock',
#     'California':'Sacramento' , 'Colorado'	:'Denver' , 'Connecticut':	'Hartford', 'Delaware'	:	'Dover',
#     'Florida':	'Tallahassee', 'Georgia':'Atlanta', 'Hawaii'	:	'Honolulu', 'Idaho'	:	'Boise',
#     'Illinois'	:	'Springfield', 'Indiana':   'Indianapolis', 'Iowa'	:	'Des Moines', 'Kansas'	:'Topeka',
#     'Kentucky'	:	'Frankfort', 'Louisiana'	:'Baton Rouge', 'Maine'	:	'Augusta', 'Maryland':'Annapolis',
#     'Massachusetts'	:'Boston', 'Michigan':'Lansing', 'Minnesota':'Saint Paul', 'Mississippi':'Jackson',
#     'Missouri':'Jefferson City', 'Montana':	'Helena', 'Nebraska':	'Lincoln', 'Nevada'	:	'Carson City',
#     'New Hampshire'	:'Concord', 'New Jersey':'Trenton', 'New Mexico':'Santa Fe', 'New York'	:'Albany',
#     'North Carolina':'Raleigh', 'North Dakota'	:'Bismarck', 'Ohio'	:'Columbus', 'Oklahoma'	:'Oklahoma City',
#     'Oregon':'Salem', 'Pennsylvania':'Harrisburg', 'Rhode Island':'Providence', 'South Carolina':'Columbia',
#     'South Dakota':'Pierre','Tennessee':'Nashville','Texas'	:'Austin', 'Utah':'Salt Lake City', 'Vermont':'Montpelier',
#     'Virginia':'Richmond', 'Washington'	:'Olympia', 'West Virginia':'Charleston', 'Wisconsin':'Madison', 'Wyoming':'Cheyenne'
# }
# for quiznum in range(35):
#     #create the quiz and answer key files.
#     quizfile = open('capitalsquiz%s.txt'%(quiznum+1),'w')
#     answerkeyfile = open('capitalsquiz_answer%s.txt' % (quiznum+1), 'w')
#
#     #write out the header for the quiz.
#     quizfile.write('Name:\n\nDate:\n\nPeriod:\n\n')
#     quizfile.write((' ' * 20) + 'State Capitals Quiz (Form %s)' % (quiznum +1))
#     quizfile.write('\n\n')
#
#     #Shuffle the order of the states.
#     states = list(capitals.keys())
#     random.shuffle(states)
#
#     for questionnum in range(50):
#
#         # Get right and wrong answers.
#         correctanswer = capitals[states[questionnum]]
#         wronganswer = list(capitals.values())
#         del wronganswer[wronganswer.index(correctanswer)]
#         wronganswer = random.sample(wronganswer,3)
#         answeroptions = wronganswer + [correctanswer]
#         random.shuffle(answeroptions)
#
#         #wrtie the question and the answer options to the quiz file.
#         quizfile.write('%s. What is the capital of %s?\n' % (questionnum +1, states[questionnum]))
#
#         for i in range(4):
#             quizfile.write('%s. %s\n' %('ABCD'[i], answeroptions[i]))
#             quizfile.write('\n')
#
#         #write the answer key to a file.
#         answerkeyfile.write('%s. %s\n' % (questionnum +1, 'ABCD'[answeroptions.index(correctanswer)]))
#     quizfile.close()
#     answerkeyfile.close()


##### 챕터 9 파일 체계화하기

# import shutil, os
# os.chdir('C:\\')

##### 챕터 10 디버깅

# raise Exception('This is the error message')
## raise, try, except로 오류 처리하기
##### (try부터 실행하고 만약에 오류나면 except로 가라)

# def boxprint(symbol, width, height):
#     if len(symbol) != 1:
#         raise Exception('Symbol must be a single character string.')
#     if width <=2 :
#         raise Exception('Width must be greater than 2.')
#     if height <=2 :
#         raise Exception('Height must be greater than 2.')
#
#     print(symbol * width)
#     for i in range(height-2):
#         print(symbol + (' ' * (width - 2)) + symbol)
#     print(symbol*width)
#
# for sym, w, h in (('*',4,4), ('0',20,5), ('x', 3,1), ('ZZ',3,3)):
#     try:
#         boxprint(sym, w, h)
#     except Exception as errb:
#         print('An exception happended: ' + str(errb))


### 역추적 내용을 문자열로 얻기

# def spam():
#     bacon()
#
# def bacon():
#     raise Exception('This is the error message')
#
# import traceback
# try:
#     raise Exception('This is the error message')
# except:
#     errorfile = open('errorInfo.txt','w')
#     errorfile.write(traceback.format_exc())
#     errorfile.close()
#     print('The traceback info was written to errorInfo.txt')


##### 단언
#
# podbaydoorstatus = 'open'
# assert podbaydoorstatus =='open', 'the pod bay doors ened to be "open".'
# podbaydoorstatus = 'I\'m sorry, Dave. I\'m afraid I can\'t do that.'
# assert podbaydoorstatus == 'open', 'the pod bay doors need to be "open".'

##### 단언 심화

# market_2nd = {'ns':'green','ew':'red'}
# mission_16th = {'ns':'red','ew': 'green'}
#
# def switchlights(stoplight):
#     for key in stoplight.keys():
#         if stoplight[key] =='green':
#             stoplight[key] = 'yellow'
#         elif stoplight[key] =='yellow':
#             stoplight[key] = 'red'
#         elif stoplight[key] == 'red':
#             stoplight[key] = 'green'
#     assert 'red' in stoplight.values(), 'Neither light is red!' + str(stoplight)
#
# print(switchlights(market_2nd))


### 로깅
#
# import logging
#
# logging.basicConfig(level = logging.DEBUG, format =' %(asctime)s - %(levelname)s - %(message)s')
# logging.debug('Start program')
#
# def factorial(n):
#     logging.debug('Start of factorial (%s)' % (n))
#     total = 1
#     for i in range(1, n+1):
#         total *= i
#         logging.debug('i is ' + str(i) + ', total is ' + str(total))
#     logging.debug('End of factorial (%s) ' % (n))
#     return total
#
# print(factorial(5))
# logging.debug('End of program')

#
# #####로깅 수준
# import logging
# logging.basicConfig(level = logging.DEBUG, format =' %(asctime)s  - %(levelname)s - %(message)s')
# logging.debug('Some debugging details.')
# logging.info('The logging module is working')
# logging.disable(logging.INFO)
# logging.debug('또 나타났다라라라라아앙')
# logging.warning('An error message is about to be logged')
# logging.error('An error has occured')
# logging.critical('The program is unable to recover')





##### 로깅 비활성화하기
## lgging.disable(logging.#그 아래로 알림끄기를 원하는 수준))


### 파이참 디버거
### 디버거를 한번 써보자

# print('Enter the first number to add:')
# first = input()
# print('Enter the second number to add:')
# second = input()
# print('Enter the third number to add:')
# third = input()
# print('The sum is ' + first + second + third)

### 디버거를 한번 더 써보자

# import random
# heads =  0
# for i in range(1,1001):
#     if random.randint(0,1) ==1:
#         heads = heads+1
#     if i == 500:
#         print('Halfway done!')
#     print('Heads came up ' + str(heads) + ' times.')

#### 11장 웹스크랩

# import webbrowser, sys, pyperclip
# # webbrowser.open('http://inventwithpython.com/')
# if len(sys.argv) >1 :
#     address = ' '.join(sys.argv[1:])
# else:
#     address = pyperclip.paste()
#
# webbrowser.open('https://www.google.com/maps/place/' + address)

### request 모듈로 웹에서 파일 다운로드하기
# import requests
# res = requests.get('http://www.gutenberg.org/cache/epub/1112/pg1112.txt')
# print(type(res))
# res.raise_for_status()
# res.status_code = requests.codes.ok
# len(res.text)
# print(res.text[:250])

### beautiful soup개체 만들기
#
# import requests, bs4
# res = requests.get('http://nostarch.com')
# res.raise_for_status()
# nostarchsoup = bs4.BeautifulSoup(res.text)
# print(type(nostarchsoup))



######### EXCEL
####################################################################################
import openpyxl
filename = "example.xlsx"
wb = openpyxl.load_workbook(filename)
print(wb.sheetnames)
sheet = wb['Sheet3']
print(type(sheet))
print(sheet.title)
anothersheet = wb.active
print(type(anothersheet))
print(anothersheet.title)
sheet=anothersheet
print(type(sheet['a1'].value))
c = sheet['b1']
print(c.value)
print('Row ' + str(c.row) + ', Column ' + str(c.column) + ' is ' + str(c.value))
print(sheet.max_row)